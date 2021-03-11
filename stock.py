# Usage
# type "python stock.py" in command line
# type in the symbol of a stock (e.g. GME)
# or type "exit" to close the program 

import openpyxl
from prettytable import PrettyTable 

# load the excel file
excelFile = openpyxl.load_workbook('example.xlsx')	# load file
stockSheet = excelFile['transaction']	# load sheet

# testData = stockSheet.cell(column=2, row=2).value
# print(testData)

inputCmd = ''


# Main
while inputCmd != 'exit':
	inputCmd = input("Please enter the ticker symbol, '-all' to show all, or 'exit' to close the program: ")

	if inputCmd == 'exit':
		break

	dateList = []
	actionList = []
	stockSymList = []
	shareList = []
	priceList = []
	totalList = []
	GLList = []

	checkStockIndicator = 0		# check if the stock is in the file (default=0, found=1)
	rowCount = 2

	if inputCmd == '-all':
		rowCount = 2
		actionCount = 0
		while True:
			if stockSheet.cell(column=3, row=rowCount).value == None:
				rowCount -= 1
				break

			stockSym = stockSheet.cell(column=3, row=rowCount).value

			dateStr = str(stockSheet.cell(column=1, row=rowCount).value)
			dateList.append(dateStr[:-9])
			actionList.append(stockSheet.cell(column=2, row=rowCount).value)
			stockSymList.append(stockSym)
			share = int(stockSheet.cell(column=4, row=rowCount).value)
			shareList.append(share)
			price = round(float(stockSheet.cell(column=5, row=rowCount).value), 2)
			priceList.append(price)
			total = round((share * price), 2)
			totalList.append(total)

			if stockSheet.cell(column=2, row=rowCount).value == 'buy':
				GLList.append('*')

			if stockSheet.cell(column=2, row=rowCount).value == 'sell':
				GLtotalShare = 0
				GLtotalCost = 0
				GLavgCost = 0
				avgGL = 0
				totalGL = 0
				checkActionCount = 2
				while True:
					if stockSheet.cell(column=3, row=checkActionCount).value == None:
						checkActionCount -= 1
						break

					if stockSheet.cell(column=3, row=checkActionCount).value == stockSym:
						if stockSheet.cell(column=2, row=checkActionCount).value == 'buy':
							stockBuyShare = int(stockSheet.cell(column=4, row=checkActionCount).value)
							stockBuyPrice = float(stockSheet.cell(column=5, row=checkActionCount).value)
							GLtotalCost += (stockBuyShare * stockBuyPrice)
							GLtotalShare += stockBuyShare
							GLavgCost = round((GLtotalCost / GLtotalShare), 2)
						elif stockSheet.cell(column=2, row=checkActionCount).value == 'sell' and checkActionCount != rowCount:
							stockSellShare = int(stockSheet.cell(column=4, row=checkActionCount).value)
							GLtotalShare -= stockSellShare
							GLtotalCost = GLtotalShare * GLavgCost
						else:
							stockSellShare = int(stockSheet.cell(column=4, row=checkActionCount).value)
							stockSellPrice = float(stockSheet.cell(column=5, row=checkActionCount).value)
							avgGL = round((stockSellPrice - GLavgCost), 2)
							totalGL = round((avgGL * stockSellShare), 2)
							GLList.append(totalGL)
					checkActionCount += 1
			actionCount += 1
			rowCount += 1

		GLshowTotal = 0
		for v in GLList:
			if v != '*':
				GLshowTotal += v
		stockTable = PrettyTable(['Date', 'Action', 'Stock', 'Share', 'Price', 'Total', 'Gain/Loss'])
		stockTable.align['Stock'] = 'l'
		stockTable.align['Share'] = 'r'
		stockTable.align['Price'] = 'r'
		stockTable.align['Total'] = 'r'
		stockTable.align['Gain/Loss'] = 'r'
		for r in range(0, actionCount):
			stockTable.add_row([dateList[r], actionList[r], stockSymList[r], shareList[r], priceList[r], totalList[r], GLList[r]])
		print("\nStock:", stockSym)
		print(stockTable)
		print("Total Gain/Loss:", GLshowTotal, '\n\n')







	while True:
		checkSymbol = stockSheet.cell(column=3, row=rowCount).value
		if checkSymbol == None:
			if checkStockIndicator == 0:
				print("Unable to find '"+inputCmd+"'. Please make sure entering in uppercase.\n")
				break
			else:
				break
		if inputCmd == checkSymbol:
			checkStockIndicator = 1
		rowCount += 1

	if checkStockIndicator == 1:
		stockSym = inputCmd
		rowCount = 2
		actionCount = 0
		while True:

			if stockSheet.cell(column=3, row=rowCount).value == None:
				rowCount -= 1
				break

			if stockSheet.cell(column=3, row=rowCount).value == stockSym:
				dateStr = str(stockSheet.cell(column=1, row=rowCount).value)
				dateList.append(dateStr[:-9])
				actionList.append(stockSheet.cell(column=2, row=rowCount).value)
				share = int(stockSheet.cell(column=4, row=rowCount).value)
				shareList.append(share)
				price = round(float(stockSheet.cell(column=5, row=rowCount).value), 2)
				priceList.append(price)
				total = round((share * price), 2)
				totalList.append(total)
				if stockSheet.cell(column=2, row=rowCount).value == 'buy':
					GLList.append('*')

				# Calculate Gain/Loss
				if stockSheet.cell(column=2, row=rowCount).value == 'sell':
					GLtotalShare = 0
					GLtotalCost = 0
					GLavgCost = 0
					avgGL = 0
					totalGL = 0
					checkActionCount = 2
					while True:
						if stockSheet.cell(column=3, row=checkActionCount).value == None:
							checkActionCount -= 1
							break

						if stockSheet.cell(column=3, row=checkActionCount).value == stockSym:
							if stockSheet.cell(column=2, row=checkActionCount).value == 'buy':
								stockBuyShare = int(stockSheet.cell(column=4, row=checkActionCount).value)
								stockBuyPrice = float(stockSheet.cell(column=5, row=checkActionCount).value)
								GLtotalCost += (stockBuyShare * stockBuyPrice)
								GLtotalShare += stockBuyShare
								GLavgCost = round((GLtotalCost / GLtotalShare), 2)
							elif stockSheet.cell(column=2, row=checkActionCount).value == 'sell' and checkActionCount != rowCount:
								stockSellShare = int(stockSheet.cell(column=4, row=checkActionCount).value)
								GLtotalShare -= stockSellShare
								GLtotalCost = GLtotalShare * GLavgCost
							else:
								stockSellShare = int(stockSheet.cell(column=4, row=checkActionCount).value)
								stockSellPrice = float(stockSheet.cell(column=5, row=checkActionCount).value)
								avgGL = round((stockSellPrice - GLavgCost), 2)
								totalGL = round((avgGL * stockSellShare), 2)
								GLList.append(totalGL)
						checkActionCount += 1




				actionCount += 1
			rowCount += 1

		if inputCmd != '-all':
			stockTable = PrettyTable(['Date', 'Action', 'Share', 'Price', 'Total', 'Gain/Loss'])
			stockTable.align['Share'] = 'r'
			stockTable.align['Price'] = 'r'
			stockTable.align['Total'] = 'r'
			stockTable.align['Gain/Loss'] = 'r'
			for r in range(0, actionCount):
				stockTable.add_row([dateList[r], actionList[r], shareList[r], priceList[r], totalList[r], GLList[r]])
			print("\nStock:", stockSym)
			print(stockTable, "\n\n")

	# print(dateList)
	# print(actionList)
	# print(shareList)
	# print(priceList)


